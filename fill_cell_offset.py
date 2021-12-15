'''
    https://qiita.com/github-nakasho/items/3f861395227e5645cce7
    https://qiita.com/github-nakasho/items/358e5602aeda81c58c81
    https://stackoverflow.com/questions/51566349/openpyxl-how-to-add-filters-to-all-columns
'''

import os
import sys
import openpyxl

output_xlsx = "output_ " + os.path.basename(sys.argv[0]) + '.xlsx'

# read input xlsx
wb1 = openpyxl.Workbook()
ws1 = wb1.worksheets[0]


green = '92d050' # R = 146, G = 208, B = 80
yello = 'ffff00' # R = 255, G = 255, B = 0

fill      = openpyxl.styles.PatternFill(patternType='solid', fgColor=yello)
fillTitle = openpyxl.styles.PatternFill(patternType='solid', fgColor=green)

side   = openpyxl.styles.borders.Side(style='thin', color='000000')
border = openpyxl.styles.borders.Border(top=side, bottom=side, left=side, right=side)

side2  = openpyxl.styles.borders.Side(style='thick', color='000000')
border2= openpyxl.styles.borders.Border(top=side, bottom=side2, left=side, right=side)

font   = openpyxl.styles.Font(name='メイリオ')

# 最初のブロック
# 絶対位置指定
for row in range(2, 10):
    for column in range(5, 10):
        cell = ws1.cell(row=row,column=column)
        if row == 2:
            cell.value = "text " + str(column)
            cell.fill = fillTitle
        else:
            cell.value = row * column
            cell.fill = fill

        cell.border = border
        cell.font = font


# 二番目のブロック
# 起点からの相対位置指定 (スクリプトの制御側の話)
offset_x = 3
offset_y = 15

for y in range(8):
    row = y + offset_y
    for x in range(5):
        column = x + offset_x
        cell = ws1.cell(row=row,column=column)
        if y == 0:
            cell.value = "text " + str(x)
            cell.fill = fillTitle
            cell.border = border2
        else:
            cell.value = x * y
            cell.fill = fill
            cell.border = border
        cell.font = font


# 三番目のブロック
# iter_rows 指定
offset_x = 5
offset_y = 25
dx = 5
dy = 15

for y, row in enumerate( ws1.iter_rows(min_row=offset_y, min_col=offset_x, max_row=offset_y + dy, max_col=offset_x + dx) ):
    for x, cell in enumerate(row):
        if y == 0:
            cell.value = "text " + str(x)
            cell.fill = fillTitle
            cell.border = border2
        else:
            cell.value = x * y
            cell.fill = fill
            cell.border = border
        cell.font = font

ws1.auto_filter.ref = ws1.dimensions

wb1.save(output_xlsx)
