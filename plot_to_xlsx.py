'''
    https://stackoverflow.com/questions/55309671/more-precise-image-placement-possible-with-openpyxl-pixel-coordinates-instead
'''

import os
import sys
import io
try:
    import openpyxl
    import numpy as np
    import matplotlib.pyplot as plt
except ImportError:
    print("pip install openpyxl numpy matplotlib")
    sys.exit(1)

from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU, pixels_to_points
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.utils import get_column_letter

c2e = cm_to_EMU
p2e = pixels_to_EMU

def cellh(x):
    """
    convert from cell height
    1 cm = 360000 EMUs
    """
    return c2e((x * 49.77)/99)

def cellw(x):
    """
    convert from cell width
    """
    return c2e((x * (18.65-1.71))/10)

def save_to_image(fig):
    """
    save fignure to image
    """
    img_data = io.BytesIO()
    fig.savefig(img_data, format='png')
    img = openpyxl.drawing.image.Image(img_data)

    # https://matplotlib.org/stable/api/figure_api.html#matplotlib.figure.Figure.figimage
    # https://www.unitconverters.net/typography/inch-to-pixel-x.htm
    img.width  = fig.get_figwidth()  * fig.dpi
    img.height = fig.get_figheight() * fig.dpi
    return img


def plot_to_excel_cell(worksheet, row, column, fig):
    """
    plot to an excel cell
    row   : 0-based
    column: 0-based
    """

    img = save_to_image(fig)
    height = img.height
    width  = img.width
    size = XDRPositiveSize2D(p2e(width), p2e(height))

    # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.cell.cell.html
    # col, row : from 1
    cell = worksheet.cell(row=row+1,column=column+1)
    width_in_font = 2.2 * int( pixels_to_points(width) / cell.font.size + 1)
    worksheet.column_dimensions[get_column_letter(column + 1)].width = width_in_font
    worksheet.row_dimensions[row+1].height   = pixels_to_points(height * 1.1)

    coloffset= cellw(0.2)
    rowoffset= cellh(0.2)

    # https://openpyxl.readthedocs.io/en/latest/api/openpyxl.drawing.spreadsheet_drawing.html
    # col, row : from 0
    marker= AnchorMarker(col=column, row=row, colOff=coloffset, rowOff=rowoffset)
    img.anchor= OneCellAnchor(_from=marker, ext=size)
    worksheet.add_image(img)

def insert_plt(worksheet, num=10, start_row=1, start_column=1):
    """
    Insert matplotlit plot to EXCEL sheet.
    """
    #green = '92d050' # R = 146, G = 208, B = 80
    #yello = 'ffff00' # R = 255, G = 255, B = 0
    #fill      = openpyxl.styles.PatternFill(patternType='solid', fgColor=yello)
    #fillTitle = openpyxl.styles.PatternFill(patternType='solid', fgColor=green)
    side   = openpyxl.styles.borders.Side(style='thin', color='000000')
    border = openpyxl.styles.borders.Border(top=side, bottom=side, left=side, right=side)

    for i in range(1, num+1):
        multiple = i + 1

        x=np.array([1,2,3,4,5])
        y=np.array([2,4,5,1,2]) * multiple

        fig, ax = plt.subplots(figsize=(4, 3))
        ax.plot(x, y)

        row    = start_row + i - 1
        column = start_column
        plot_to_excel_cell(worksheet, row, column, fig)

        cell = worksheet.cell(row=row+1,column=column)
        cell.border = border

        cell = worksheet.cell(row=row+1,column=column+1)
        cell.border = border


output_xlsx = "output_ " + os.path.basename(sys.argv[0]) + '.xlsx'

wb1 = openpyxl.Workbook()
ws1 = wb1.worksheets[0]
insert_plt(ws1)

wb1.save(output_xlsx)
