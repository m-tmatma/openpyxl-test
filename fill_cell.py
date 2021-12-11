'''
    https://qiita.com/github-nakasho/items/3f861395227e5645cce7
    https://qiita.com/github-nakasho/items/358e5602aeda81c58c81
    https://stackoverflow.com/questions/51566349/openpyxl-how-to-add-filters-to-all-columns
'''

import openpyxl
import os
import sys

output_xlsx = "output_ " + os.path.basename(sys.argv[0]) + '.xlsx'

from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font

# read input xlsx
wb1 = openpyxl.Workbook()
ws1 = wb1.worksheets[0]


green = '92d050' # R = 146, G = 208, B = 80
yello = 'ffff00' # R = 255, G = 255, B = 0

fill      = PatternFill(patternType='solid', fgColor=yello)
fillTitle = PatternFill(patternType='solid', fgColor=green)

side = Side(style='thin', color='000000')
border = Border(top=side, bottom=side, left=side, right=side)
font = Font(name='メイリオ')

for row in range(2, 10):
    for column in range(5, 10):
        cell = ws1.cell(row=row,column=column)
        if row == 2:
            cell.value = "text " + str(column)
        else:
            cell.value = row * column

for i, row in enumerate(ws1):
    for cell in row:
        if i == 0:
            ws1[cell.coordinate].fill = fillTitle
        else:
            ws1[cell.coordinate].fill = fill
        ws1[cell.coordinate].border = border
        ws1[cell.coordinate].font = font

ws1.auto_filter.ref = ws1.dimensions

wb1.save(output_xlsx)
