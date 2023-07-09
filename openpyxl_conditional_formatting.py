'''
Test Program for conditional_formatting
'''
from openpyxl import Workbook
#from openpyxl.styles import Color, Font, Border
from openpyxl.styles import PatternFill, GradientFill
from openpyxl.styles import Alignment
#from openpyxl.styles.differential import DifferentialStyle
#from openpyxl.formatting.rule import ColorScaleRule, Rule
from openpyxl.formatting.rule import FormulaRule
from openpyxl.formatting.rule import CellIsRule

wb = Workbook()
ws = wb.active

redFill = PatternFill(start_color='EE1111', end_color='EE1111', fill_type='solid')
gadientFill = GradientFill(type='linear', stop=("00FF00", "FF0000"))

ws.conditional_formatting.add('D1:D10', FormulaRule(formula=['NOT(ISBLANK(D1))'], stopIfTrue=True, fill=redFill))
ws.conditional_formatting.add('E1:E10', FormulaRule(formula=['ISBLANK(E1)'], stopIfTrue=True, fill=redFill))
ws.conditional_formatting.add('F1:F10', FormulaRule(formula=['F1=1'], stopIfTrue=True, fill=gadientFill))
ws.conditional_formatting.add('G1:G10', CellIsRule(operator='lessThan', formula=[5], stopIfTrue=True, fill=redFill))

ws['D1'] = "A"
ws['D2'] = "B"
ws['D3'] = 10.2
ws['D3'].alignment = Alignment(horizontal='right') 

ws['E1'] = "A"
ws['E2'] = "B"

ws['F1'] = 0
ws['F2'] = 1

for i in range(1, 10+1):
	ws[f'G{i}'] = i

wb.save("test.xlsx")
